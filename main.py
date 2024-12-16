from openpyxl import load_workbook
from docx import Document
from functions_EXCEL import *
from functions_WORD import * 
import os
import pathlib
from win32com.client import Dispatch
import sys
import time
import pathlib
import re

def excluirImagensPATH(imagem):
    if os.path.exists(imagem):
        os.remove(imagem)
    print("✔ -> Removido imagem da pasta raiz do projeto!")

def pegarGraficosExcel(app, workbook_file_name, workbook):
    app.DisplayAlerts = False

    for i, sheet in enumerate(workbook.Worksheets):
        for chartObject in sheet.ChartObjects():
            chartObject.Chart.Export(rf"{str(pathlib.Path().resolve())}\chart{str(i+1)}.png")
            i +=1
    workbook.Close(SaveChanges=False, Filename=workbook_file_name)
    print("✔ -> Obter gráficos do arquivo EXCEL!")

## Inicialização do app =========
print(r"""
________            _____________                           __________  
___  __ \_________________  /__(_)______ _______________    ___  ___  \ 
__  /_/ /_  ___/  _ \  __  /__  /__  __ `__ \  _ \  ___/    __  / _ \  |
_  ____/_  /   /  __/ /_/ / _  / _  / / / / /  __/ /__      _  / , _/ / 
/_/     /_/    \___/\__,_/  /_/  /_/ /_/ /_/\___/\___/      | /_/|_| /  
                                                             \______/   
Iniciando processo de formatação...
      """)

arquivocorreto = False
while arquivocorreto == False:
    padraoArquivo = "<>:\"/|?*"
    nomeArquivoUser = str(input("\n💾 -> Digite o nome do arquivo a ser salvo: "))
    for i in padraoArquivo:
        if i in nomeArquivoUser:
            print("\n⛔ -> O nome do arquivo NÃO pode ter os seguintes caracteres: '/' , '<', '>', '\', '|', '?', '*'.")
            arquivocorreto = False
            break
        else: 
            arquivocorreto = True
try: 
    arquivos = os.listdir('./')
    for arquivo in arquivos:
        if arquivo.endswith(".docx"):
            ARQUIVO_WORD = arquivo
        if arquivo.endswith(".xlsm") or arquivo.endswith(".xlsx"):
            ARQUIVO_EXCEL = arquivo
    f = open(ARQUIVO_WORD, 'rb')
    g = open(ARQUIVO_EXCEL, 'rb')
    documentoWord = Document(f)
    ws = load_workbook(g, data_only=True)
except:
    print("ERRO: Erro ao identificar arquivos para formatação.")
    time.sleep(10)
    sys.exit(1)
print("\n✔ -> Verificar arquivos WORD e EXCEL!")

planilha = ws["Listagem"]
WORD_criarTabelaListagem(planilha, documentoWord)
WORD_criarTabelasOS(documentoWord)

if len(planilhas_OS) != 0:
    datacorreta = False
    padrao = r"^\d{2}/\d{2}/\d{4}$"
    while not datacorreta:
        data = str(input("\n📆 -> Digite a data a ser colocada nas OS (dd/mm/aaaa): "))
        datacorreta = bool(re.match(padrao, data))
        if not datacorreta:
            print("\n⛔ -> A data deve ter o formato dd/mm/aaaa, é só aceito números e barras '/'")
    WORD_addValoresTabelaOS(ws, documentoWord, data)


PASTA_RESULTADOS = "RELATÓRIOS FORMATADOS"
os.makedirs(PASTA_RESULTADOS, exist_ok=True)
caminhoWord = os.path.join(PASTA_RESULTADOS, f"{nomeArquivoUser}.docx")        

try:
    app = Dispatch("Excel.Application")
    workbook_file_name = rf"{str(pathlib.Path().resolve())}\{ARQUIVO_EXCEL}"
    workbook = app.Workbooks.Open(Filename=workbook_file_name)
    pegarGraficosExcel(app, workbook_file_name, workbook)

    for i in documentoWord.paragraphs:
        if i.text == "[grafico_status]":
            WORD_addGraficos(i, 2)
            continue
        elif i.text == "[grafico_tendencia]":
            WORD_addGraficos(i, 4)
except:
    print("ERRO: Erro ao inserir imagens dos gráficos do arquivo WORD.")
    excluirImagensPATH("chart2.png")
    excluirImagensPATH("chart3.png")
    excluirImagensPATH("chart4.png")             
    time.sleep(10)
    sys.exit(1)
documentoWord.save(caminhoWord)

excluirImagensPATH("chart2.png")
excluirImagensPATH("chart3.png")
excluirImagensPATH("chart4.png")             

print("\nArquivos formatados com sucesso!\n")
time.sleep(10)
